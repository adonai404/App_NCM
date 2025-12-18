import streamlit as st
import pandas as pd
import os
import openpyxl
import tempfile
from pathlib import Path

def processar_planilha(planilha_ncm, pasta_saida, data, descricao, imposto, vinculo_credito, base_credito, planilha_modelo):
    try:
        # 1. Processar planilha NCM
        st.info("Processando planilha NCM...")
        progress_bar = st.progress(0)
        
        # Lê a planilha Excel começando da linha 6
        df = pd.read_excel(planilha_ncm, dtype=str, header=5)
        df = df.fillna('')
        df['NCM'] = df['NCM'].str.replace('.', '')
        
        # Cria pasta temporária para arquivos TXT
        pasta_temp = os.path.join(pasta_saida, "temp")
        os.makedirs(pasta_temp, exist_ok=True)
        
        # Processa cada CST único
        total_cst = len(df['CST PIS/COFINS ENTRADA'].unique())
        for idx, cst in enumerate(df['CST PIS/COFINS ENTRADA'].unique()):
            df_cst = df[df['CST PIS/COFINS ENTRADA'] == cst]
            
            # Define o tipo do CST
            tipo_cst = {
                '73': 'ALIQUOTA ZERO',
                '70': 'MONOFÁSICO',
                '50': 'TRIBUTADO'
            }.get(cst, f'CST {cst}')
            
            # Processa cada natureza única para o CST atual
            for natureza in df_cst['CÓDIGO SPED'].unique():
                df_nat = df_cst[df_cst['CÓDIGO SPED'] == natureza]
                ncms = list(set(df_nat['NCM'].tolist()))
                ncms.sort()
                cst_saida = df_nat['CST PIS/COFINS SAÍDA'].iloc[0] if not df_nat.empty else ''
                
                # Cria o nome do arquivo
                nome_arquivo = f"{tipo_cst} - {natureza}.txt"
                caminho_arquivo = os.path.join(pasta_temp, nome_arquivo)
                
                # Prepara o conteúdo do arquivo
                conteudo = f"CST PIS/COFINS ENTRADA: {cst}\n"
                conteudo += f"CST PIS/COFINS SAÍDA: {cst_saida}\n"
                conteudo += f"NATUREZA: {natureza}\n\n"
                conteudo += "NCMs:\n"
                conteudo += "\n".join(ncms)
                
                # Salva o arquivo
                with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                    f.write(conteudo)
            
            # Atualiza progresso
            progress_bar.progress((idx + 1) / total_cst * 0.3)
        
        st.info("Criando planilhas individuais...")
        
        # 2. Criar planilhas individuais
        arquivos_txt = [f for f in os.listdir(pasta_temp) if f.endswith('.txt')]
        total_arquivos = len(arquivos_txt)
        
        for idx, arquivo_txt in enumerate(arquivos_txt, 1):
            caminho_txt = os.path.join(pasta_temp, arquivo_txt)
            
            # Analisa o arquivo TXT
            with open(caminho_txt, 'r', encoding='utf-8') as f:
                linhas = f.readlines()
            
            # Extrai informações
            cst_entrada = ""
            cst_saida = ""
            natureza = ""
            ncms = []
            
            for linha in linhas:
                linha = linha.strip()
                if linha.startswith("CST PIS/COFINS ENTRADA:"):
                    cst_entrada = linha.split(":", 1)[1].strip()
                elif linha.startswith("CST PIS/COFINS SAÍDA:"):
                    cst_saida = linha.split(":", 1)[1].strip()
                elif linha.startswith("NATUREZA:"):
                    natureza = linha.split(":", 1)[1].strip()
                elif linha == "NCMs:":
                    continue
                elif linha and not linha.startswith("CST") and not linha.startswith("NATUREZA"):
                    ncms.append(linha)
            
            # Cria planilha individual
            wb = openpyxl.load_workbook(planilha_modelo)
            ws = wb.active
            
            # Preenche campos
            ws['B2'] = data
            ws['C2'] = descricao
            ws['B4'] = os.path.splitext(arquivo_txt)[0]
            ws['B6'] = imposto
            ws['C6'] = cst_entrada
            ws['D6'] = vinculo_credito
            ws['E6'] = base_credito
            ws['H6'] = cst_saida
            ws['I6'] = natureza
            
            # Preenche NCMs
            for i, ncm in enumerate(ncms):
                ws.cell(row=8 + i, column=1, value='NCM')
                ws.cell(row=8 + i, column=2, value=ncm)
            
            # Salva planilha
            nome_saida = os.path.splitext(arquivo_txt)[0] + ".xlsx"
            caminho_saida = os.path.join(pasta_saida, nome_saida)
            wb.save(caminho_saida)
            
            # Atualiza progresso
            progress_bar.progress(0.3 + (idx / total_arquivos * 0.4))
        
        st.info("Mesclando planilhas...")
        
        # 3. Mesclar planilhas
        wb_final = openpyxl.load_workbook(planilha_modelo)
        ws_final = wb_final.active
        
        # Define onde começar a inserir os dados
        current_row = 3
        
        # Para cada planilha individual
        arquivos_excel = [f for f in os.listdir(pasta_saida) if f.endswith('.xlsx') and f != "planilha_final.xlsx"]
        total_excel = len(arquivos_excel)
        
        for idx, arquivo in enumerate(arquivos_excel, 1):
            caminho_arquivo = os.path.join(pasta_saida, arquivo)
            
            # Carrega planilha
            wb_data = openpyxl.load_workbook(caminho_arquivo)
            ws_data = wb_data.active
            
            # Encontra última linha com dados
            last_row = 0
            for r in range(ws_data.max_row, 0, -1):
                if any(ws_data.cell(r, c).value is not None for c in range(1, ws_data.max_column+1)):
                    last_row = r
                    break
            
            if last_row >= 3:
                # Copia dados
                for r in range(3, last_row + 1):
                    for c in range(1, ws_data.max_column + 1):
                        ws_final.cell(row=current_row, column=c, value=ws_data.cell(row=r, column=c).value)
                    current_row += 1
            
            # Atualiza progresso
            progress_bar.progress(0.7 + (idx / total_excel * 0.3))
        
        # Salva planilha final
        caminho_final = os.path.join(pasta_saida, "planilha_final.xlsx")
        wb_final.save(caminho_final)
        
        # Limpa arquivos temporários
        for arquivo in os.listdir(pasta_temp):
            os.remove(os.path.join(pasta_temp, arquivo))
        os.rmdir(pasta_temp)
        
        progress_bar.progress(1.0)
        st.success("Processamento concluído!")
        return caminho_final
        
    except Exception as e:
        st.error(f"Erro durante o processamento: {str(e)}")
        return None

def main():
    st.set_page_config(
        page_title="Processador de NCM",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("Processador de NCM")
    st.markdown("---")
    
    # Sidebar para configurações
    with st.sidebar:
        st.header("Configurações")
        data = st.date_input("Data")
        descricao = st.text_input("Descrição")
        imposto = st.selectbox("Imposto", ['C', 'N', 'T', 'S'])
        
        # Toggle para campos de crédito
        mostrar_campos_credito = st.toggle("Mostrar campos de crédito", value=False)
        
        if mostrar_campos_credito:
            vinculo_credito = st.text_input("Vínculo do Crédito")
            base_credito = st.text_input("Base de Crédito")
        else:
            vinculo_credito = ""
            base_credito = ""
    
    # Área principal
    st.subheader("Arquivos de Entrada")
    planilha_modelo = st.file_uploader("Planilha Modelo", type=['xlsx'])
    planilha_ncm = st.file_uploader("Planilha com NCMs", type=['xlsx'])
    
    st.markdown("---")
    
    # Botão processar
    if st.button("Processar", type="primary"):
        if not planilha_modelo or not planilha_ncm:
            st.error("Selecione a planilha modelo e a planilha com NCMs")
            return
        
        # Cria pasta temporária para saída
        pasta_saida = os.path.join(tempfile.gettempdir(), "ncm_output")
        os.makedirs(pasta_saida, exist_ok=True)
        
        # Processa os arquivos
        caminho_final = processar_planilha(
            planilha_ncm,
            pasta_saida,
            data.strftime("%d/%m/%Y"),
            descricao,
            imposto,
            vinculo_credito,
            base_credito,
            planilha_modelo
        )
        
        if caminho_final:
            st.success(f"Arquivo final salvo em: {caminho_final}")
            
            # Botão para download
            with open(caminho_final, "rb") as f:
                st.download_button(
                    label="Baixar Planilha Final",
                    data=f,
                    file_name="planilha_final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

if __name__ == "__main__":
    main() 
